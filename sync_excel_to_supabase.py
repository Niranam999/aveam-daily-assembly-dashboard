import os
import json
import re
import sys
import urllib.request
import urllib.parse
from urllib.error import HTTPError, URLError

# Set encoding to UTF-8
sys.stdout.reconfigure(encoding='utf-8')

# Paths to files
DIRECTORY = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(DIRECTORY, 'supabase_config.json')
F5_PATH = 'D:\\AVEAM\\Antigravity2\\Test-LiveArtifact\\5_PN_Job_Relationship_rev00-test.xlsx'
F1_PATH = os.path.join(DIRECTORY, '1-List of each sub-assembly.xlsx')

def load_or_create_config():
    """Load config file or prompt user to create it if missing."""
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            try:
                return json.load(f)
            except Exception as e:
                print(f"[Error] Failed to parse config file: {e}")
                
    print("\n==================================================")
    print("      AVEAM SUPABASE SETUP CONFIGURATION")
    print("==================================================")
    url = input("กรุณาใส่ Supabase Project URL: ").strip()
    key = input("กรุณาใส่ Supabase Anon Key หรือ Service Role Key: ").strip()
    gas_url = input("กรุณาใส่ Google Apps Script Web App URL (ถ้ามี, หรือกด Enter ข้าม): ").strip()
    
    config = {
        "supabase_url": url,
        "supabase_key": key,
        "google_sheet_url": gas_url
    }
    
    with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
        json.dump(config, f, indent=4)
        
    print(f"บันทึกไฟล์ตั้งค่าสำเร็จที่: {CONFIG_PATH}\n")
    return config

# Load config
config = load_or_create_config()
SUPABASE_URL = config.get("supabase_url")
SUPABASE_KEY = config.get("supabase_key")

if not SUPABASE_URL or not SUPABASE_KEY:
    print("[Error] Supabase URL หรือ Key ใน supabase_config.json ไม่ถูกต้อง")
    sys.exit(1)

# Ensure URLs are clean
if SUPABASE_URL.endswith('/'):
    SUPABASE_URL = SUPABASE_URL[:-1]

def make_supabase_request(endpoint, method='GET', data=None, headers=None):
    """Make HTTP requests to the Supabase REST API."""
    url = f"{SUPABASE_URL}/rest/v1/{endpoint}"
    
    default_headers = {
        'apikey': SUPABASE_KEY,
        'Authorization': f"Bearer {SUPABASE_KEY}",
        'Content-Type': 'application/json'
    }
    if headers:
        default_headers.update(headers)
        
    req_data = None
    if data:
        req_data = json.dumps(data).encode('utf-8')
        
    req = urllib.request.Request(url, data=req_data, headers=default_headers, method=method)
    
    try:
        with urllib.request.urlopen(req) as response:
            res_content = response.read().decode('utf-8')
            if res_content:
                return json.loads(res_content)
            return True
    except HTTPError as e:
        error_msg = e.read().decode('utf-8')
        print(f"[Supabase API Error] {method} {endpoint} failed with code {e.code}: {error_msg}")
        raise e
    except URLError as e:
        print(f"[Supabase Network Error] {e.reason}")
        raise e

# ----------------------------------------------------
# Parse Excel files (adapted from update_prototype_data.py)
# ----------------------------------------------------
def clean_thai_desc(proc_id, name):
    name_str = str(name).strip()
    manual_map = {
        "parts receiving": "รับพาร์ท",
        "receiving material": "รับพาร์ท",
        "parts check": "รับพาร์ท",
        "parts check (ตรวจสอบพาร์ทก่อนพ่นสี)": "ตรวจสี",
        "stamping pn (แสตมป์พาร์ทนัมเบอร์)": "แสตมป์พาร์ท",
        "etching pn (เอจชิ่งพาร์ทนัมเบอร์)": "เอจชิ่งพาร์ท",
        "pin installation-base plate (ติดตั้งพินที่แผ่น base plate)": "ติดพิน Base",
        "pin installation-top plate (ติดตั้งพินที่แผ่น top plate)": "ติดพิน Top",
        "pin installation-side plate (ติดตั้งพินที่แผ่น top plate)": "ติดพิน Side",
        "top plate assy (ประกอบชุดท็อปเพลท)": "ประกอบท็อป",
        "handle assy (ประกอบชุดมือจับ)": "ประกอบมือจับ",
        "ground cable assy (ประกอบสายกราวน์)": "ประกอบสายดิน",
        "land yard cable assy (ประกอบสายแลนยาร์ด)": "ประกอบแลนยาร์ด",
        "main assy (ประกอบชุดเมนแอสเซมบลี)": "ประกอบหลัก",
        "final inspection (การตรวจสอบขั้นตอนสุดท้าย)": "ตรวจสเต็ปสุดท้าย",
        "packing (การแพ็คงาน)": "แพ็คงาน",
        "storing": "จัดเก็บ",
        "main pn 01-21-24466": "ประกอบหลัก",
        "sub assy pn 01-21-24455": "ประกอบย่อย",
        "schematic 19-21-24510": "เดินสายไฟ",
        "function check/test": "ทดสอบเครื่อง",
        "final inspection": "ตรวจงาน",
        "storing in wh": "จัดเก็บเข้าคลัง"
    }
    
    key = name_str.lower().replace("  ", " ").strip()
    if key in manual_map:
        return manual_map[key]
        
    parentheses = re.findall(r'\((.*?)\)', name_str)
    for p in parentheses:
        if any('\u0e00' <= char <= '\u0e7f' for char in p):
            th_text = p.strip()
            th_text = th_text.replace("แสตมป์พาร์ทนัมเบอร์", "แสตมป์พาร์ท").replace("เอจชิ่งพาร์ทนัมเบอร์", "เอจชิ่งพาร์ท")
            th_text = th_text.replace("ติดตั้งพินที่แผ่น", "ติดพิน").replace("ประกอบชุด", "ประกอบ")
            th_text = th_text.replace("การตรวจสอบขั้นตอนสุดท้าย", "ตรวจสเต็ปสุดท้าย")
            return th_text
            
    if any('\u0e00' <= char <= '\u0e7f' for char in name_str):
        th_parts = re.findall(r'[\u0e00-\u0e7f\s\d]+', name_str)
        th_text = "".join(th_parts).strip()
        th_text = re.sub(r'\s+', ' ', th_text)
        if th_text:
            return th_text
            
    if "receiving" in key: return "รับพาร์ท"
    if "inspect" in key: return "ตรวจงาน"
    if "packing" in key: return "แพ็คงาน"
    if "cable" in key:
        if "CABLE-" in proc_id.upper():
            return f"สายไฟ {proc_id.upper().replace('CABLE-', '')}"
        return "สายไฟ"
    if "wiring" in key: return "เดินสายไฟ"
    if "testing" in key: return "ทดสอบเครื่อง"
    if "main-assy" in key or "main assy" in key: return "ประกอบหลัก"
    if "storing" in key or "wh" in key: return "จัดเก็บ"
    
    return ""

def load_excel_data():
    import openpyxl
    print("[1/3] กำลังโหลดข้อมูลจากไฟล์ Excel...")
    
    if not os.path.exists(F5_PATH):
        print(f"[Error] ไม่พบไฟล์ที่ {F5_PATH}")
        sys.exit(1)
    if not os.path.exists(F1_PATH):
        print(f"[Error] ไม่พบไฟล์ที่ {F1_PATH}")
        sys.exit(1)
        
    # Load sub-assembly details
    wb1 = openpyxl.load_workbook(F1_PATH, data_only=True)
    sh1 = wb1['Project and Sub-Assembly']
    master_sub_assemblies = {}
    current_proj = None
    
    for r in range(3, sh1.max_row+1):
        proj_code = sh1.cell(r, 3).value
        if proj_code:
            current_proj = str(proj_code).strip()
            master_sub_assemblies[current_proj] = []
            
        proc = sh1.cell(r, 7).value
        proc_desc = sh1.cell(r, 8).value
        
        if current_proj and (proc is not None or proc_desc is not None):
            pn, desc = ("", str(proc_desc).strip()) if ";" not in str(proc_desc) else str(proc_desc).split(";", 1)
            pn = pn.strip()
            desc = desc.strip()
            master_sub_assemblies[current_proj].append({
                'process_id': str(proc).strip() if proc is not None else "",
                'pn': pn,
                'name': desc
            })
            
    # Load master times
    wb5 = openpyxl.load_workbook(F5_PATH, data_only=True)
    sh_master = wb5['Std Assembly Time']
    master_info = {}
    for r in range(3, sh_master.max_row+1):
        pn = sh_master.cell(r, 4).value
        if pn:
            pn_clean = str(pn).strip()
            master_info[pn_clean] = {
                'customer': str(sh_master.cell(r, 2).value or '').strip(),
                'project_code': str(sh_master.cell(r, 3).value or '').strip(),
                'description': str(sh_master.cell(r, 5).value or '').strip(),
                'std_time': float(sh_master.cell(r, 6).value or 0.0)
            }
            
    # Load active WIP jobs
    sh_jobs = wb5['PN_Job_Relationship']
    headers = [sh_jobs.cell(1, c).value for c in range(1, sh_jobs.max_column+1)]
    
    c_cust = headers.index('Customer') + 1
    c_so = headers.index('SO No.') + 1
    c_proj = headers.index('Project Code') + 1
    c_pn = headers.index('Part Number') + 1
    c_desc = headers.index('Description') + 1
    c_qty = headers.index('Qty') + 1
    c_progress = headers.index('Job Progress (%)') + 1
    c_eval_progress = headers.index('My Evaluate Progress (%)') + 1
    c_tl = headers.index('TEAM LEADER') + 1
    c_backlog = headers.index('BACKLOG/UNASSIGNED') + 1
    c_assigned = headers.index('ASSIGNED') + 1
    c_in_progress = headers.index('IN PROGRESS') + 1
    c_qa = headers.index('QA/INSPECTION') + 1
    c_complete = headers.index('COMPLETE') + 1
    c_mc = headers.index('MC Number') + 1
    c_job = headers.index('Job Number') + 1
    c_m1 = headers.index('MEMBER 1') + 1
    
    projects_grouped = {}
    
    for r in range(2, sh_jobs.max_row+1):
        val_y = sh_jobs.cell(r, c_backlog).value
        val_z = sh_jobs.cell(r, c_assigned).value
        val_aa = sh_jobs.cell(r, c_in_progress).value
        val_ab = sh_jobs.cell(r, c_qa).value
        val_ac = sh_jobs.cell(r, c_complete).value
        
        is_active = any(v == '✓' for v in [val_y, val_z, val_aa, val_ab, val_ac])
        if not is_active:
            continue
            
        cust = sh_jobs.cell(r, c_cust).value
        proj = sh_jobs.cell(r, c_proj).value
        so = sh_jobs.cell(r, c_so).value
        pn = sh_jobs.cell(r, c_pn).value
        desc = sh_jobs.cell(r, c_desc).value
        qty = sh_jobs.cell(r, c_qty).value
        prog = sh_jobs.cell(r, c_progress).value
        eval_prog = sh_jobs.cell(r, c_eval_progress).value
        tl = sh_jobs.cell(r, c_tl).value
        
        if not cust or not pn:
            continue
            
        cust_str = str(cust).strip()
        proj_str = str(proj).strip() if proj else 'PRX'
        if cust_str == 'ULC':
            proj_str = 'PRX'
            
        so_str = str(int(float(so))) if isinstance(so, (int, float)) else (str(so).strip() if so else '99999')
        qty_val = int(float(qty)) if isinstance(qty, (int, float)) else 1
        
        prog_val = 0.0
        if prog is not None:
            try:
                prog_val = float(prog)
                if prog_val <= 1.0 and prog_val > 0.0: prog_val *= 100.0
            except: pass
        elif eval_prog is not None:
            try:
                prog_val = float(eval_prog)
                if prog_val <= 1.0 and prog_val > 0.0: prog_val *= 100.0
            except: pass
                
        tl_str = str(tl).strip() if tl else 'WANLOP CHANPHET'
        mc_str = str(sh_jobs.cell(r, c_mc).value or '').strip()
        m1_str = str(sh_jobs.cell(r, c_m1).value or '').strip()
        job_val = sh_jobs.cell(r, c_job).value
        job_str = str(int(float(job_val))) if isinstance(job_val, (int, float)) else (str(job_val).strip() if job_val else '')
        
        # Determine this row's kanban stage
        kanban_stage = 'backlog'
        if val_ac == '✓':
            kanban_stage = 'completed'
        elif val_ab == '✓':
            kanban_stage = 'qa_inspection'
        elif val_aa == '✓':
            kanban_stage = 'in_progress'
        elif val_z == '✓':
            kanban_stage = 'assigned'
            
        if cust_str == 'ULC':
            group_key = (cust_str, proj_str, 'ULC_ALL')
        else:
            group_key = (cust_str, proj_str, job_str)
            
        if group_key not in projects_grouped:
            projects_grouped[group_key] = {
                'customer': cust_str,
                'project_code': proj_str,
                'so': so_str,
                'qty': qty_val,
                'team_leader': tl_str,
                'member_1': m1_str,
                'mc_number': mc_str,
                'jobs': []
            }
        else:
            projects_grouped[group_key]['qty'] += qty_val
            if mc_str and (not projects_grouped[group_key]['mc_number'] or projects_grouped[group_key]['mc_number'] == '-'):
                projects_grouped[group_key]['mc_number'] = mc_str
            if m1_str and (not projects_grouped[group_key]['member_1'] or projects_grouped[group_key]['member_1'] == '-'):
                projects_grouped[group_key]['member_1'] = m1_str
            
        # Find standard time for this row/part
        row_std_time = 0.0
        pn_clean = str(pn).strip()
        if pn_clean in master_info:
            row_std_time = master_info[pn_clean]['std_time']
        else:
            for mpn, m in master_info.items():
                if mpn.upper() == pn_clean.upper():
                    row_std_time = m['std_time']
                    break

        projects_grouped[group_key]['jobs'].append({
            'pn': pn_clean,
            'name': str(desc).strip() if desc else '',
            'progress': int(prog_val),
            'job_no': job_str,
            'qty': qty_val,
            'std_time': row_std_time,
            'kanban_stage': kanban_stage
        })
        
    excel_projects = []
    for key, data in projects_grouped.items():
        cust_str, proj_str, group_id_str = key
        
        if cust_str == 'ULC':
            job_no_display = ""
            proj_id = "ULC-PRX-ALL"
        else:
            job_no_display = group_id_str
            proj_id = f"{cust_str}-{proj_str}-{job_no_display.replace(' ', '')}"
            
        master_proj_code = next((p for p in master_sub_assemblies.keys() if p.upper() == proj_str.upper()), None)
        
        subs = []
        if master_proj_code and len(master_sub_assemblies[master_proj_code]) > 0:
            for ms in master_sub_assemblies[master_proj_code]:
                prog_vals = []
                for job in data['jobs']:
                    if job['pn'] == ms['pn'] or (ms['pn'] and ms['pn'] in job['pn']) or (job['pn'] and job['pn'] in ms['pn']):
                        prog_vals.append(job['progress'])
                
                if prog_vals:
                    prog_val = sum(prog_vals) / len(prog_vals)
                else:
                    prog_val = 0
                    
                if ms['process_id'] == 'R-1' and any(j['progress'] > 0 for j in data['jobs']):
                    prog_val = 100
                if ms['process_id'] == 'MAIN-ASSY' and prog_val == 0:
                    main_prog_vals = []
                    for job in data['jobs']:
                        if proj_str.upper() in job['name'].upper() or 'ASSY' in job['name'].upper():
                            main_prog_vals.append(job['progress'])
                    if main_prog_vals:
                        prog_val = sum(main_prog_vals) / len(main_prog_vals)
                            
                th_desc = clean_thai_desc(ms['process_id'], ms['name'])
                subs.append({
                    'pn': ms['process_id'],
                    'name': f"{ms['pn']}; {ms['name']}" if ms['pn'] else ms['name'],
                    'th_desc': th_desc,
                    'progress': int(prog_val)
                })
        else:
            for job in data['jobs']:
                subs.append({
                    'pn': job['pn'],
                    'name': job['name'],
                    'th_desc': '',
                    'progress': job['progress'],
                    'qty': job.get('qty', 0),
                    'std_time': job.get('std_time', 0.0)
                })
                
        overall_prog = 0
        main_row_prog = next((j['progress'] for j in data['jobs'] if j['pn'] in master_info), None)
        if main_row_prog is not None:
            overall_prog = main_row_prog
        else:
            overall_prog = sum(s['progress'] for s in subs) / len(subs) if subs else 0
            
        est_hours = 0.0
        main_desc = proj_str
        for job in data['jobs']:
            if job['pn'] in master_info:
                est_hours = master_info[job['pn']]['std_time']
                main_desc = master_info[job['pn']]['description']
                break
        if est_hours == 0.0:
            for mpn, m in master_info.items():
                if m['project_code'].upper() == proj_str.upper():
                    est_hours = m['std_time']
                    main_desc = m['description']
                    break
        if cust_str == 'ULC':
            # Calculate average std_time of all jobs in this ULC group
            ulc_times = []
            for job in data['jobs']:
                if job['pn'] in master_info:
                    ulc_times.append(master_info[job['pn']]['std_time'])
                else:
                    found = False
                    for mpn, m in master_info.items():
                        if mpn.upper() == job['pn'].upper():
                            ulc_times.append(m['std_time'])
                            found = True
                            break
                    if not found:
                        ulc_times.append(0.0)
            
            if ulc_times:
                est_hours = round(sum(ulc_times) / len(ulc_times), 4)
            else:
                est_hours = 1.1292
                
            main_desc = 'PRX250 Standard Module Assembly'
            
        # Aggregate Kanban stage counts
        kanban_backlog = sum(1 for j in data['jobs'] if j.get('kanban_stage') == 'backlog')
        kanban_assigned = sum(1 for j in data['jobs'] if j.get('kanban_stage') == 'assigned')
        kanban_in_progress = sum(1 for j in data['jobs'] if j.get('kanban_stage') == 'in_progress')
        kanban_qa = sum(1 for j in data['jobs'] if j.get('kanban_stage') == 'qa_inspection')
        kanban_completed = sum(1 for j in data['jobs'] if j.get('kanban_stage') == 'completed')

        main_pn = '-'
        if cust_str == 'ULC':
            main_pn = 'PRX250-All Assy'
        elif data['jobs']:
            main_pn = next((j['pn'] for j in data['jobs'] if j['pn'] in master_info), data['jobs'][0]['pn'])
            
        excel_projects.append({
            'id': proj_id,
            'customer': cust_str,
            'project_code': proj_str,
            'part_number': main_pn,
            'description': main_desc,
            'status': 'completed' if (overall_prog >= 100 or kanban_completed > 0) else 'ontime',
            'team_leader': data['team_leader'],
            'member_1': data['member_1'] if data['member_1'] else '-',
            'job_no': job_no_display,
            'mc_number': data['mc_number'] if data['mc_number'] else '-',
            'progress': int(overall_prog),
            'qty': data['qty'],
            'qty_done': 0, # Initialized to 0, merged later
            'est_hours': est_hours,
            'sub_assemblies': subs,
            'kanban_backlog': kanban_backlog,
            'kanban_assigned': kanban_assigned,
            'kanban_in_progress': kanban_in_progress,
            'kanban_qa': kanban_qa,
            'kanban_completed': kanban_completed
        })
        
    print(f"โหลดข้อมูลสำเร็จ: พบ {len(excel_projects)} จ๊อบที่มีสถานะ WIP ใน Excel")
    return excel_projects

# ----------------------------------------------------
# Sync with Supabase (HTTP-based postgrest REST API)
# ----------------------------------------------------
def sync_to_supabase(excel_projects):
    print("[2/3] กำลังดึงสถานะโครงการปัจจุบันจาก Supabase...")
    
    # 1. Fetch existing projects to merge progress
    try:
        existing_projects = make_supabase_request('projects?select=id,progress,qty_done,sub_assemblies')
        existing_dict = {p['id']: p for p in existing_projects}
        print(f"พบโครงการในฐานข้อมูลคลาวด์ {len(existing_projects)} รายการ")
    except Exception as e:
        print("[Error] ดึงข้อมูลจาก Supabase ล้มเหลว โปรดตรวจสอบ URL หรือ API Key ใน supabase_config.json")
        sys.exit(1)
        
    # 2. Merge progress
    merged_projects = []
    for ep in excel_projects:
        pid = ep['id']
        if pid in existing_dict:
            # Preserve current progress & cumulative finished goods from cloud database
            db_proj = existing_dict[pid]
            ep['progress'] = db_proj['progress']
            ep['qty_done'] = db_proj['qty_done']
            
            # Map sub-assembly progress
            db_subs = db_proj.get('sub_assemblies', [])
            if db_subs:
                db_sub_dict = {s['pn']: s['progress'] for s in db_subs if 'pn' in s}
                for sub in ep['sub_assemblies']:
                    if sub['pn'] in db_sub_dict:
                        sub['progress'] = db_sub_dict[sub['pn']]
                        
            # Recalculate status based on merged progress
            if ep['progress'] >= 100:
                ep['status'] = 'completed'
            # Keep existing status if it was changed to 'delayed'
            elif db_proj.get('status') == 'delayed':
                ep['status'] = 'delayed'
                
        merged_projects.append(ep)
        
    # 3. Upload merged projects to Supabase (Upsert via POST with Prefer header)
    print("[3/3] กำลังอัปโหลดข้อมูลรวมไปยัง Supabase...")
    try:
        # PostgREST bulk upsert (merge-duplicates based on primary key 'id')
        make_supabase_request(
            'projects',
            method='POST',
            data=merged_projects,
            headers={'Prefer': 'resolution=merge-duplicates'}
        )
        print("อัปโหลดเรียบร้อย! ข้อมูลโครงการทั้งหมดตรงกันกับ Excel และคงความคืบหน้าของพนักงาน")
    except Exception as e:
        print(f"[Error] อัปโหลดข้อมูลโครงการล้มเหลว: {e}")
        sys.exit(1)
        
    # 4. Clean up inactive jobs
    active_ids = {p['id'] for p in excel_projects}
    inactive_ids = [pid for pid in existing_dict.keys() if pid not in active_ids]
    
    if inactive_ids:
        print(f"กำลังลบโครงการที่ประกอบเสร็จ/ไม่อยู่ในแผน WIP จำนวน {len(inactive_ids)} รายการ...")
        try:
            # Delete inactive projects
            encoded_ids = ",".join([urllib.parse.quote(pid) for pid in inactive_ids])
            make_supabase_request(f"projects?id=in.({encoded_ids})", method='DELETE')
            print("ลบรายการที่ไม่มีสถานะ WIP เรียบร้อย")
        except Exception as e:
            print(f"[Warning] ลบโครงการเก่าล้มเหลว: {e}")

if __name__ == '__main__':
    print("==================================================")
    print("      AVEAM EXCEL SYNC TO SUPABASE RUNNING")
    print("==================================================")
    
    try:
        excel_projects = load_excel_data()
        sync_to_supabase(excel_projects)
        print("\n[SUCCESS] การซิงค์เสร็จสิ้นสมบูรณ์! แดชบอร์ดออนไลน์อัปเดตแล้ว")
    except KeyboardInterrupt:
        print("\nยกเลิกการซิงค์ข้อมูล")
    except Exception as e:
        print(f"\n[FAIL] มีข้อผิดพลาดเกิดขึ้นระหว่างการประมวลผล: {e}")
    print("==================================================")
