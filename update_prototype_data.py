import openpyxl
import json
import os
import sys
import re

# Reconfigure stdout to use UTF-8
sys.stdout.reconfigure(encoding='utf-8')

f5_path = 'D:\\AVEAM\\Antigravity2\\Test-LiveArtifact\\5_PN_Job_Relationship_rev00-test.xlsx'
f1_path = 'd:\\AVEAM\\Antigravity2\\Assembly Progress Dailly Report\\1-List of each sub-assembly.xlsx'
out_js_path = 'd:\\AVEAM\\Antigravity2\\Assembly Progress Dailly Report\\data.js'

if not os.path.exists(f5_path):
    print(f"Error: {f5_path} not found.")
    sys.exit(1)

if not os.path.exists(f1_path):
    print(f"Error: {f1_path} not found.")
    sys.exit(1)

# Helper to extract or translate to clean Thai descriptions
def clean_thai_desc(proc_id, name):
    name_str = str(name).strip()
    
    manual_map = {
        "parts receiving": "รับพาร์ท",
        "receiving material": "รับพาร์ท",
        "parts check": "รับพาร์ท",
        "parts check (ตรวจสอบพาร์ทก่อนพ่นสี)": "ตรวจสี",
        "stamping pn (แสตมป์พาร์ทนัมเบอร์)": "แสตมป์พาร์ท",
        "etching pn (เอจชิ่งพาร์ทนัมเบอร์)": "เอจชิ่งพาร์ท",
        "pin installation-base plate (ติดตั้งพินที่แผ่น base plate)": "ติดพิน Base",
        "pin installation-top plate (ติดตั้งพินที่แผ่น top plate)": "ติดพิน Top",
        "pin installation-side plate (ติดตั้งพินที่แผ่น top plate)": "ติดพิน Side",
        "top plate assy (ประกอบชุดท็อปเพลท)": "ประกอบท็อป",
        "handle assy (ประกอบชุดมือจับ)": "ประกอบมือจับ",
        "ground cable assy (ประกอบสายกราวน์)": "ประกอบสายดิน",
        "land yard cable assy (ประกอบสายแลนยาร์ด)": "ประกอบแลนยาร์ด",
        "main assy (ประกอบชุดเมนแอสเซมบลี)": "ประกอบหลัก",
        "final inspection (การตรวจสอบขั้นตอนสุดท้าย)": "ตรวจสเต็ปสุดท้าย",
        "packing (การแพ็คงาน)": "แพ็คงาน",
        "storing": "จัดเก็บ",
        "main pn 01-21-24466": "ประกอบหลัก",
        "sub assy pn 01-21-24455": "ประกอบย่อย",
        "schematic 19-21-24510": "เดินสายไฟ",
        "function check/test": "ทดสอบเครื่อง",
        "final inspection": "ตรวจงาน",
        "storing in wh": "จัดเก็บเข้าคลัง"
    }
    
    key = name_str.lower().replace("  ", " ").strip()
    if key in manual_map:
        return manual_map[key]
        
    # Check if there is text inside parentheses ( ... )
    parentheses = re.findall(r'\((.*?)\)', name_str)
    for p in parentheses:
        # Check if contains Thai characters (\u0e00-\u0e7f)
        if any('\u0e00' <= char <= '\u0e7f' for char in p):
            th_text = p.strip()
            th_text = th_text.replace("แสตมป์พาร์ทนัมเบอร์", "แสตมป์พาร์ท")
            th_text = th_text.replace("เอจชิ่งพาร์ทนัมเบอร์", "เอจชิ่งพาร์ท")
            th_text = th_text.replace("ติดตั้งพินที่แผ่น", "ติดพิน")
            th_text = th_text.replace("ประกอบชุด", "ประกอบ")
            th_text = th_text.replace("การตรวจสอบขั้นตอนสุดท้าย", "ตรวจสเต็ปสุดท้าย")
            return th_text
            
    # Check if there are any Thai characters in the name
    if any('\u0e00' <= char <= '\u0e7f' for char in name_str):
        th_parts = re.findall(r'[\u0e00-\u0e7f\s\d]+', name_str)
        th_text = "".join(th_parts).strip()
        th_text = re.sub(r'\s+', ' ', th_text)
        if th_text:
            return th_text
            
    # Fallback for common English terms
    if "receiving" in key: return "รับพาร์ท"
    if "inspect" in key: return "ตรวจงาน"
    if "packing" in key: return "แพ็คงาน"
    if "cable" in key:
        if "CABLE-" in proc_id.upper():
            c_num = proc_id.upper().replace("CABLE-", "")
            return f"สายไฟ {c_num}"
        return "สายไฟ"
    if "wiring" in key: return "เดินสายไฟ"
    if "testing" in key: return "ทดสอบเครื่อง"
    if "main-assy" in key or "main assy" in key: return "ประกอบหลัก"
    if "storing" in key or "wh" in key: return "จัดเก็บ"
    
    return ""

# 1. Load sub-assemblies (processes) from 1-List of each sub-assembly.xlsx columns G and H
wb1 = openpyxl.load_workbook(f1_path, data_only=True)
sh1 = wb1['Project and Sub-Assembly']

master_sub_assemblies = {}
current_proj = None

for r in range(3, sh1.max_row+1):
    proj_code = sh1.cell(r, 3).value
    if proj_code:
        current_proj = str(proj_code).strip()
        master_sub_assemblies[current_proj] = []
        
    proc = sh1.cell(r, 7).value
    proc_desc = sh1.cell(r, 8).value
    
    if current_proj and (proc is not None or proc_desc is not None):
        pn, desc = parse_process_desc = ("", str(proc_desc).strip()) if ";" not in str(proc_desc) else str(proc_desc).split(";", 1)
        pn = pn.strip()
        desc = desc.strip()
        master_sub_assemblies[current_proj].append({
            'process_id': str(proc).strip() if proc is not None else "",
            'pn': pn,
            'name': desc
        })

print(f"Loaded sub-assembly definitions for {len(master_sub_assemblies)} projects from File 1.")

# 2. Load Master standard times from File 5
wb5 = openpyxl.load_workbook(f5_path, data_only=True)
sh_master = wb5['Std Assembly Time']
master_info = {}
for r in range(3, sh_master.max_row+1):
    cust = sh_master.cell(r, 2).value
    proj_code = sh_master.cell(r, 3).value
    pn = sh_master.cell(r, 4).value
    desc = sh_master.cell(r, 5).value
    std_time = sh_master.cell(r, 6).value
    if pn:
        pn_clean = str(pn).strip()
        master_info[pn_clean] = {
            'customer': str(cust).strip() if cust else '',
            'project_code': str(proj_code).strip() if proj_code else '',
            'description': str(desc).strip() if desc else '',
            'std_time': float(std_time) if std_time is not None else 0.0
        }

# 3. Load Jobs from File 5
sh_jobs = wb5['PN_Job_Relationship']
headers = [sh_jobs.cell(1, c).value for c in range(1, sh_jobs.max_column+1)]

# Find column indexes
c_cust = headers.index('Customer') + 1
c_so = headers.index('SO No.') + 1
c_po = headers.index('PO No.') + 1
c_proj = headers.index('Project Code') + 1
c_pn = headers.index('Part Number') + 1
c_desc = headers.index('Description') + 1
c_qty = headers.index('Qty') + 1
c_progress = headers.index('Job Progress (%)') + 1
c_eval_progress = headers.index('My Evaluate Progress (%)') + 1
c_tl = headers.index('TEAM LEADER') + 1
c_in_progress = headers.index('IN PROGRESS') + 1
c_mc = headers.index('MC Number') + 1
c_job = headers.index('Job Number') + 1
c_m1 = headers.index('MEMBER 1') + 1

# Group jobs by (Customer, Project Code, SO No.) where IN PROGRESS == '✓'
projects_grouped = {}

for r in range(2, sh_jobs.max_row+1):
    in_progress = sh_jobs.cell(r, c_in_progress).value
    if in_progress != '✓':
        continue
        
    cust = sh_jobs.cell(r, c_cust).value
    proj = sh_jobs.cell(r, c_proj).value
    so = sh_jobs.cell(r, c_so).value
    pn = sh_jobs.cell(r, c_pn).value
    desc = sh_jobs.cell(r, c_desc).value
    qty = sh_jobs.cell(r, c_qty).value
    prog = sh_jobs.cell(r, c_progress).value
    eval_prog = sh_jobs.cell(r, c_eval_progress).value
    tl = sh_jobs.cell(r, c_tl).value
    
    if not cust or not pn:
        continue
        
    cust_str = str(cust).strip()
    proj_str = str(proj).strip() if proj else 'PRX'
    if cust_str == 'ULC':
        proj_str = 'PRX'
        
    so_str = str(int(float(so))) if isinstance(so, (int, float)) else (str(so).strip() if so else '99999')
    qty_val = int(float(qty)) if isinstance(qty, (int, float)) else 1
    
    # Progress parsing
    prog_val = 0.0
    if prog is not None:
        try:
            prog_val = float(prog)
            if prog_val <= 1.0 and prog_val > 0.0:
                prog_val *= 100.0
        except:
            pass
    elif eval_prog is not None:
        try:
            prog_val = float(eval_prog)
            if prog_val <= 1.0 and prog_val > 0.0:
                prog_val *= 100.0
        except:
            pass
            
    tl_str = str(tl).strip() if tl else 'WANLOP CHANPHET'
    mc = sh_jobs.cell(r, c_mc).value
    mc_str = str(mc).strip() if mc else ''
    m1 = sh_jobs.cell(r, c_m1).value
    m1_str = str(m1).strip() if m1 else ''
    
    # Get Job Number (Col L)
    job_val = sh_jobs.cell(r, c_job).value
    job_str = str(int(float(job_val))) if isinstance(job_val, (int, float)) else (str(job_val).strip() if job_val else '')
    
    # ULC projects are grouped under a single key to show overall system progress as a single card
    if cust_str == 'ULC':
        group_key = (cust_str, proj_str, 'ULC_ALL')
    else:
        # Others are grouped by their specific Job Number (Column L) so they appear as separate cards
        group_key = (cust_str, proj_str, job_str)
        
    if group_key not in projects_grouped:
        projects_grouped[group_key] = {
            'customer': cust_str,
            'project_code': proj_str,
            'so': so_str,
            'qty': qty_val,
            'team_leader': tl_str,
            'member_1': m1_str,
            'mc_number': mc_str,
            'jobs': []
        }
    else:
        projects_grouped[group_key]['qty'] += qty_val
        if mc_str and (not projects_grouped[group_key]['mc_number'] or projects_grouped[group_key]['mc_number'] == '-'):
            projects_grouped[group_key]['mc_number'] = mc_str
        if m1_str and (not projects_grouped[group_key]['member_1'] or projects_grouped[group_key]['member_1'] == '-'):
            projects_grouped[group_key]['member_1'] = m1_str
        
    # Find standard time for this row/part
    row_std_time = 0.0
    pn_clean = str(pn).strip()
    if pn_clean in master_info:
        row_std_time = master_info[pn_clean]['std_time']
    else:
        for mpn, m in master_info.items():
            if mpn.upper() == pn_clean.upper():
                row_std_time = m['std_time']
                break

    projects_grouped[group_key]['jobs'].append({
        'pn': pn_clean,
        'name': str(desc).strip() if desc else '',
        'progress': int(prog_val),
        'job_no': job_str,
        'qty': qty_val,
        'std_time': row_std_time
    })

# Format projects for prototype data.js
formatted_projects = []
for key, data in projects_grouped.items():
    cust_str, proj_str, group_id_str = key
    
    # Determine the Job Number display format
    if cust_str == 'ULC':
        job_no_display = ''
        so_str = data['so']
    else:
        job_no_display = group_id_str
        so_str = data['so']
    
    # Check if this project has sub-assemblies in the master list
    master_proj_code = None
    for pcode in master_sub_assemblies.keys():
        if pcode.upper() == proj_str.upper():
            master_proj_code = pcode
            break
            
    subs = []
    if master_proj_code and len(master_sub_assemblies[master_proj_code]) > 0:
        # Load sub-assemblies defined in File 1 columns G and H
        master_subs = master_sub_assemblies[master_proj_code]
        for ms in master_subs:
            # Match progress:
            # If there's an active job row in File 5 matching this sub-assembly's part number
            prog_val = 0
            for job in data['jobs']:
                if job['pn'] == ms['pn'] or (ms['pn'] and ms['pn'] in job['pn']) or (job['pn'] and job['pn'] in ms['pn']):
                    prog_val = job['progress']
                    break
                    
            # Fallback: if progress is 0, but the overall project has progress, mark R-1 (Receiving material) as 100%
            if ms['process_id'] == 'R-1' and any(j['progress'] > 0 for j in data['jobs']):
                prog_val = 100
                
            # If progress is 0 and it's MAIN-ASSY, use the main job progress
            if ms['process_id'] == 'MAIN-ASSY' and prog_val == 0:
                for job in data['jobs']:
                    if proj_str.upper() in job['name'].upper() or 'ASSY' in job['name'].upper():
                        prog_val = job['progress']
                        break
            
            # Clean and get Thai description
            th_desc = clean_thai_desc(ms['process_id'], ms['name'])
            
            subs.append({
                'pn': ms['process_id'],
                'name': f"{ms['pn']}; {ms['name']}" if ms['pn'] else ms['name'],
                'th_desc': th_desc,
                'progress': prog_val
            })
    else:
        # Fallback to active job rows from File 5 (like ULC, UIC)
        for job in data['jobs']:
            subs.append({
                'pn': job['pn'],
                'name': job['name'],
                'th_desc': '', # Fallback doesn't need th_desc since the name itself is already english/thai
                'progress': job['progress'],
                'qty': job.get('qty', 0),
                'std_time': job.get('std_time', 0.0)
            })
            
    # Calculate overall progress
    overall_prog = 0
    main_row_prog = None
    for job in data['jobs']:
        # If the job PN matches the project code or the main assembly PN in master standard times
        if job['pn'] in master_info:
            main_row_prog = job['progress']
            break
    if main_row_prog is not None:
        overall_prog = main_row_prog
    else:
        overall_prog = sum(s['progress'] for s in subs) / len(subs) if subs else 0
        
    # Find main description and est_hours from master list
    est_hours = 0.0
    main_desc = proj_str
    
    for job in data['jobs']:
        if job['pn'] in master_info:
            m = master_info[job['pn']]
            est_hours = m['std_time']
            main_desc = m['description']
            break
            
    if est_hours == 0.0:
        for mpn, m in master_info.items():
            if m['project_code'].upper() == proj_str.upper():
                est_hours = m['std_time']
                main_desc = m['description']
                break
                
    if cust_str == 'ULC':
        # Calculate average std_time of all jobs in this ULC group
        ulc_times = []
        for job in data['jobs']:
            if job['pn'] in master_info:
                ulc_times.append(master_info[job['pn']]['std_time'])
            else:
                found = False
                for mpn, m in master_info.items():
                    if mpn.upper() == job['pn'].upper():
                        ulc_times.append(m['std_time'])
                        found = True
                        break
                if not found:
                    ulc_times.append(0.0)
        
        if ulc_times:
            est_hours = round(sum(ulc_times) / len(ulc_times), 4)
        else:
            est_hours = 1.1292
            
        main_desc = 'PRX250 Standard Module Assembly'
        
    # Find main part number
    main_pn = '-'
    if cust_str == 'ULC':
        main_pn = 'PRX250-All Assy'
    elif data['jobs']:
        for job in data['jobs']:
            if job['pn'] in master_info:
                main_pn = job['pn']
                break
        if main_pn == '-':
            main_pn = data['jobs'][0]['pn']

    status = 'ontime'
    if overall_prog >= 100:
        status = 'completed'
        
    proj_id = f"{proj_str}-{job_no_display.replace(' ', '')}"
    if cust_str == 'ULC':
        proj_id = "ULC-PRX-ALL"
        
    formatted_projects.append({
        'id': proj_id,
        'customer': cust_str,
        'project_code': proj_str,
        'part_number': main_pn,
        'description': main_desc,
        'status': status,
        'team_leader': data['team_leader'],
        'member_1': data['member_1'] if data['member_1'] else '-',
        'job_no': job_no_display,
        'mc_number': data['mc_number'] if data['mc_number'] else '-',
        'progress': int(overall_prog),
        'qty': data['qty'],
        'est_hours': est_hours,
        'sub_assemblies': subs
    })

# 4. Write data.js
with open(out_js_path, 'w', encoding='utf-8') as f:
    f.write("// Live data compiled from 5_PN_Job_Relationship_rev00-test.xlsx & 1-List of each sub-assembly.xlsx\n")
    f.write("const EXCEL_PROJECTS = ")
    json.dump(formatted_projects, f, ensure_ascii=False, indent=4)
    f.write(";\n")

print(f"Successfully compiled {len(formatted_projects)} active projects with sub-assemblies to {out_js_path}!")
