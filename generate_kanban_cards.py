import os
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
import urllib.request
import urllib.parse

# Paths Configuration
DIRECTORY = os.path.dirname(os.path.abspath(__file__))
F5_PATH = r"D:\AVEAM\Antigravity2\Test-LiveArtifact\5_PN_Job_Relationship_rev00-test.xlsx"
F5_DIR = os.path.dirname(F5_PATH)
KANBAN_PATH = os.path.join(F5_DIR, "8-Kanban card.xlsx")
TEMP_QR_DIR = os.path.join(DIRECTORY, "temp_qrs")

# Create temporary directory for QR codes
os.makedirs(TEMP_QR_DIR, exist_ok=True)

# Custom Thai console output encoding
sys.stdout.reconfigure(encoding='utf-8')

def clean_value(val):
    """Clean Excel values into strings, removing floating point decimals for IDs."""
    if val is None:
        return ""
    val_str = str(val).strip()
    if val_str.endswith(".0"):
        val_str = val_str[:-2]
    return val_str

def draw_card_outer_border(ws, start_row, start_col, end_row, end_col):
    """Draws a thick dark outline border around each card."""
    medium = Side(border_style="medium", color="1C1E21")
    
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            # Retain existing cell borders if any
            t = medium if r == start_row else Side(border_style=None)
            b = medium if r == end_row else Side(border_style=None)
            l = medium if c == start_col else Side(border_style=None)
            rg = medium if c == end_col else Side(border_style=None)
            cell.border = Border(top=t, bottom=b, left=l, right=rg)

def main():
    print("\n==========================================================")
    print("      ระบบสร้างการ์ดคันบังติดงานผลิตอัตโนมัติ (Kanban Generator)")
    print("==========================================================")
    
    if not os.path.exists(F5_PATH):
        print(f"[Error] ไม่พบไฟล์หลักที่ {F5_PATH}")
        sys.exit(1)
        
    print(f"กำลังสแกนจ๊อบที่กำลังทำงานและจัดแผนงานจาก {os.path.basename(F5_PATH)}...")
    wb = openpyxl.load_workbook(F5_PATH, data_only=True)
    ws_src = wb["PN_Job_Relationship"]
    
    headers = [ws_src.cell(row=1, column=c).value for c in range(1, ws_src.max_column + 1)]
    
    try:
        c_cust = headers.index("Customer") + 1
        c_proj = headers.index("Project Code") + 1
        c_pn = headers.index("Part Number") + 1
        c_desc = headers.index("Description") + 1
        c_mc = headers.index("MC Number") + 1
        c_proj_id = headers.index("ProjectID") + 1
        c_job = headers.index("Job Number") + 1
        c_qty = headers.index("Qty") + 1
        c_assigned = headers.index("ASSIGNED") + 1
        c_in_progress = headers.index("IN PROGRESS") + 1
        c_tl = headers.index("TEAM LEADER") + 1
        c_m1 = headers.index("MEMBER 1") + 1
    except ValueError as e:
        print(f"[Error] คอลัมน์ที่จำเป็นไม่ครบถ้วนในไฟล์หลัก: {e}")
        sys.exit(1)
        
    active_jobs = []
    
    # Read active jobs (marked in ASSIGNED or IN PROGRESS)
    for r in range(2, ws_src.max_row + 1):
        assigned_val = clean_value(ws_src.cell(r, c_assigned).value)
        progress_val = clean_value(ws_src.cell(r, c_in_progress).value)
        
        # Check if row is assigned or in progress (marked with checkmark)
        if "✓" in assigned_val or "✓" in progress_val:
            cust = clean_value(ws_src.cell(r, c_cust).value)
            proj = clean_value(ws_src.cell(r, c_proj).value)
            pn = clean_value(ws_src.cell(r, c_pn).value)
            desc = clean_value(ws_src.cell(r, c_desc).value)
            mc = clean_value(ws_src.cell(r, c_mc).value)
            proj_id = clean_value(ws_src.cell(r, c_proj_id).value)
            job = clean_value(ws_src.cell(r, c_job).value)
            qty = clean_value(ws_src.cell(r, c_qty).value)
            tl = clean_value(ws_src.cell(r, c_tl).value)
            m1 = clean_value(ws_src.cell(r, c_m1).value)
            
            # Combine other members if available (Columns 32 to 35)
            members = [m1] if m1 else []
            for col_m in range(32, 36):
                m_val = clean_value(ws_src.cell(r, col_m).value)
                if m_val:
                    members.append(m_val)
            
            active_jobs.append({
                "customer": cust,
                "project_code": proj,
                "part_number": pn,
                "description": desc,
                "mc_number": mc if mc else "ไม่มีข้อมูล",
                "project_id": proj_id if proj_id else "ไม่มีข้อมูล",
                "job_number": job,
                "qty": qty,
                "team_leader": tl,
                "members": ", ".join(members) if members else "ไม่มีข้อมูลสมาชิก"
            })
            
    print(f"สแกนเสร็จสิ้น! พบจ๊อบที่ต้องสร้างการ์ดคันบังทั้งหมด: {len(active_jobs)} รายการ")
    
    if not active_jobs:
        print("[Notice] ไม่มีจ๊อบที่เปิดสถานะ Assigned หรือ In Progress จึงไม่มีการสร้างการ์ด")
        return
        
    # --- CREATE KANBAN WORKBOOK ---
    wb_k = openpyxl.Workbook()
    ws_k = wb_k.active
    ws_k.title = "Kanban Cards"
    
    # Enable grid lines
    ws_k.views.sheetView[0].showGridLines = True
    
    # Page setup for A4 printing
    ws_k.page_setup.orientation = ws_k.ORIENTATION_PORTRAIT
    ws_k.page_setup.paperSize = ws_k.PAPERSIZE_A4
    ws_k.page_margins.left = 0.25
    ws_k.page_margins.right = 0.25
    ws_k.page_margins.top = 0.25
    ws_k.page_margins.bottom = 0.25
    
    # Define Column Widths (A4 optimized layout)
    ws_k.column_dimensions['A'].width = 3
    ws_k.column_dimensions['B'].width = 12
    ws_k.column_dimensions['C'].width = 16
    ws_k.column_dimensions['D'].width = 12
    ws_k.column_dimensions['E'].width = 16
    ws_k.column_dimensions['F'].width = 3
    ws_k.column_dimensions['G'].width = 12
    ws_k.column_dimensions['H'].width = 16
    ws_k.column_dimensions['I'].width = 12
    ws_k.column_dimensions['J'].width = 16
    ws_k.column_dimensions['K'].width = 3
    
    # Custom Brand Colors for Card Headers (Depending on Customer)
    customer_colors = {
        "VEECO": "10B981",    # Emerald Green
        "COHU": "8B5CF6",     # Royal Purple
        "UIC": "3B82F6",      # Bright Blue
        "ULC": "F59E0B",      # Warm Orange
        "DEFAULT": "475569"   # Slate Grey
    }
    
    # Define Styles
    f_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    f_label = Font(name="Segoe UI", size=9, bold=True, color="555555")
    f_value = Font(name="Segoe UI", size=9, color="000000")
    f_bold_val = Font(name="Segoe UI", size=10, bold=True, color="000000")
    f_banner = Font(name="Segoe UI", size=8, italic=True, color="64748B")
    
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")
    
    fill_banner = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_bg = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # Border definitions
    thin_border = Border(
        bottom=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0')
    )

    total_pages = (len(active_jobs) + 7) // 8
    print(f"กำลังจัดเตรียมหน้ากระดาษ A4 ทั้งหมด: {total_pages} หน้า...")
    
    # Loop to generate cards
    for idx, job in enumerate(active_jobs):
        page = idx // 8
        slot = idx % 8
        row_idx = slot // 2
        col_idx = slot % 2
        
        # Calculate cell coordinates
        # 9 content rows per card + 1 spacer row = 10 rows per card slot -> 40 rows per A4 page
        page_offset = page * 41
        start_row = page_offset + row_idx * 10 + 2
        start_col = 2 if col_idx == 0 else 7 # Column B or Column G
        
        # Setup row heights for the page if not set
        if slot == 0:
            ws_k.row_dimensions[page_offset + 1].height = 15 # Top margin
            
        ws_k.row_dimensions[start_row].height = 24     # Header
        ws_k.row_dimensions[start_row + 1].height = 18 # Data 1 (Job No)
        ws_k.row_dimensions[start_row + 2].height = 18 # Data 2 (Part No)
        ws_k.row_dimensions[start_row + 3].height = 18 # Data 3 (Machine No)
        ws_k.row_dimensions[start_row + 4].height = 18 # Data 4 (ProjectID)
        ws_k.row_dimensions[start_row + 5].height = 18 # Data 5 (Qty)
        ws_k.row_dimensions[start_row + 6].height = 18 # Data 6 (Lead)
        ws_k.row_dimensions[start_row + 7].height = 18 # Data 7 (Members / แม่ทีม)
        ws_k.row_dimensions[start_row + 8].height = 14 # Banner
        ws_k.row_dimensions[start_row + 9].height = 15 # Spacer row
        
        # Determine Header Color
        cust_upper = job["customer"].upper()
        h_color = customer_colors.get(cust_upper, customer_colors["DEFAULT"])
        fill_header = PatternFill(start_color=h_color, end_color=h_color, fill_type="solid")
        
        # Fill base background cells
        for r in range(start_row, start_row + 9):
            for c in range(start_col, start_col + 4):
                ws_k.cell(row=r, column=c).fill = fill_bg
                
        # 1. Card Header (Merged row)
        ws_k.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 3)
        h_cell = ws_k.cell(row=start_row, column=start_col)
        h_cell.value = f" {job['customer']} - {job['project_code']}"
        h_cell.font = f_header
        h_cell.alignment = align_center
        h_cell.fill = fill_header
        
        # 2. Card Content (Left Columns)
        labels = [
            ("Job No:", job["job_number"], f_bold_val),
            ("Part No:", job["part_number"], f_value),
            ("Machine No:", job["mc_number"], f_value),
            ("ProjectID:", job["project_id"], f_value),
            ("Qty:", f"{job['qty']} Units" if job['qty'] else "ไม่มีข้อมูล", f_bold_val),
            ("Lead:", job["team_leader"], f_value),
            ("แม่ทีม:", job["members"], f_value)
        ]
        
        for offset, (label, val, val_font) in enumerate(labels):
            curr_row = start_row + 1 + offset
            
            l_cell = ws_k.cell(row=curr_row, column=start_col)
            l_cell.value = label
            l_cell.font = f_label
            l_cell.alignment = align_right
            
            v_cell = ws_k.cell(row=curr_row, column=start_col + 1)
            v_cell.value = val
            v_cell.font = val_font
            v_cell.alignment = align_left
            
            # Draw thin borders internally
            l_cell.border = thin_border
            v_cell.border = thin_border
            ws_k.cell(row=curr_row, column=start_col + 2).border = thin_border
            ws_k.cell(row=curr_row, column=start_col + 3).border = thin_border
            
        # 3. Card Bottom Banner (Merged scan reminder)
        ws_k.merge_cells(start_row=start_row + 8, start_column=start_col, end_row=start_row + 8, end_column=start_col + 3)
        b_cell = ws_k.cell(row=start_row + 8, column=start_col)
        b_cell.value = "📲 Scan QR to Update Assembly Dashboard"
        b_cell.font = f_banner
        b_cell.alignment = align_center
        b_cell.fill = fill_banner
        b_cell.border = thin_border
        
        # 4. Generate & Insert QR Code
        # URL template: https://niranam999.github.io/aveam-daily-assembly-dashboard/update.html?job=JOB_NUMBER
        update_url = f"https://niranam999.github.io/aveam-daily-assembly-dashboard/update.html?job={job['job_number']}"
        qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=115x115&data={urllib.parse.quote(update_url)}"
        
        qr_file = os.path.join(TEMP_QR_DIR, f"qr_{job['job_number']}.png")
        
        try:
            # Download QR code image from API
            urllib.request.urlretrieve(qr_url, qr_file)
            
            # Place image starting at column D/I (c+2) and row 2 of the card
            img = Image(qr_file)
            img_cell = ws_k.cell(row=start_row + 1, column=start_col + 2)
            ws_k.add_image(img, img_cell.coordinate)
        except Exception as e:
            print(f"[Warning] ไม่สามารถโหลด QR code สำหรับจ๊อบ {job['job_number']}: {e}")
            
        # 5. Draw Card Outer Outline Border (Medium Border)
        draw_card_outer_border(ws_k, start_row, start_col, start_row + 8, start_col + 3)
        
    # Save Workbook
    try:
        wb_k.save(KANBAN_PATH)
        print(f"\n[Success] การ์ดคันบังถูกสร้างและจัดเก็บในไฟล์เรียบร้อยแล้ว!")
        print(f"  - คลังตำแหน่งไฟล์: {KANBAN_PATH}")
    except Exception as e:
        print(f"[Error] ไม่สามารถเซฟไฟล์การ์ดคันบังได้ (กรุณาเช็กว่าเปิดไฟล์ค้างไว้หรือไม่): {e}")
        sys.exit(1)
        
    # --- CLEANUP TEMP QR IMAGES ---
    try:
        for f in os.listdir(TEMP_QR_DIR):
            os.remove(os.path.join(TEMP_QR_DIR, f))
        os.rmdir(TEMP_QR_DIR)
    except:
        pass

if __name__ == '__main__':
    main()
