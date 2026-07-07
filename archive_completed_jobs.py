import os
import sys
import copy
import re
import shutil
from datetime import datetime

# Paths configuration
F5_PATH = 'D:\\AVEAM\\Antigravity2\\Test-LiveArtifact\\5_PN_Job_Relationship_rev00-test.xlsx'
F5_DIR = os.path.dirname(F5_PATH)
COMPLETED_PATH = os.path.join(F5_DIR, '7_5_PN_Job_Relationship_COMPLTED_rev00.xlsx')

# Set console encoding to UTF-8 to prevent Thai unicode display errors
sys.stdout.reconfigure(encoding='utf-8')

def adjust_formula_row(formula_str, src_row, dest_row):
    """Adjusts row references inside Excel formulas when cells are moved."""
    if not isinstance(formula_str, str) or not formula_str.startswith('='):
        return formula_str
    # Preceded by column letters (A-Z, AA-AZ) and not followed by a digit
    pattern = r'([A-Z]+)' + str(src_row) + r'\b'
    adjusted = re.sub(pattern, r'\g<1>' + str(dest_row), formula_str)
    return adjusted

def get_date_key(val):
    """Converts the Latest Ship Date to a datetime object for sorting."""
    if val is None:
        return datetime.max  # Push empty dates to the bottom
    if isinstance(val, datetime):
        return val
    
    val_str = str(val).strip()
    # Try common formats
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y"):
        try:
            return datetime.strptime(val_str, fmt)
        except ValueError:
            continue
            
    # If the string contains date characters but fails to parse, extract digits
    date_match = re.search(r'(\d{4})[-/](\d{1,2})[-/](\d{1,2})', val_str)
    if date_match:
        try:
            return datetime(int(date_match.group(1)), int(date_match.group(2)), int(date_match.group(3)))
        except ValueError:
            pass
            
    date_match_reverse = re.search(r'(\d{1,2})[-/](\d{1,2})[-/](\d{4})', val_str)
    if date_match_reverse:
        try:
            return datetime(int(date_match_reverse.group(3)), int(date_match_reverse.group(2)), int(date_match_reverse.group(1)))
        except ValueError:
            pass

    return datetime.max

def get_row_data(sh_val, sh_form, r):
    """Reads a row's values, data types, and formatting styles."""
    cells = []
    for col in range(1, sh_form.max_column + 1):
        cell = sh_form.cell(row=r, column=col)
        cells.append({
            'value': cell.value,
            'data_type': cell.data_type,
            'font': copy.copy(cell.font),
            'fill': copy.copy(cell.fill),
            'border': copy.copy(cell.border),
            'alignment': copy.copy(cell.alignment),
            'number_format': cell.number_format
        })
        
    val_q = sh_val.cell(r, 17).value   # Q (Col 17) - Evaluated Date
    val_u = sh_val.cell(r, 21).value   # U (Col 21) - Evaluated Completed Status
    val_ac = sh_val.cell(r, 29).value  # AC (Col 29) - Evaluated Complete Checkbox
    
    # Row is completed if U is 'Completed' or AC is '✓'
    is_completed = (val_u == 'Completed' or val_ac == '✓')
    
    return {
        'original_row_index': r,
        'cells': cells,
        'sort_val': val_q,
        'is_completed': is_completed
    }

def clear_rows(sheet):
    """Deletes all rows starting from row 2 downwards."""
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row)

def write_rows(sheet, rows_data):
    """Writes compiled rows back to the sheet, copying styles and adjusting formulas."""
    for idx, row in enumerate(rows_data):
        dest_row = idx + 2  # Start writing from row 2 (leaving row 1 header)
        src_row = row['original_row_index']
        
        for col_idx, cell_data in enumerate(row['cells']):
            col = col_idx + 1
            cell = sheet.cell(row=dest_row, column=col)
            
            # Write value (update formulas if the row index changes)
            val = cell_data['value']
            if isinstance(val, str) and val.startswith('='):
                val = adjust_formula_row(val, src_row, dest_row)
                
            cell.value = val
            cell.data_type = cell_data['data_type']
            
            # Copy style properties
            if cell_data['font']: cell.font = cell_data['font']
            if cell_data['fill']: cell.fill = cell_data['fill']
            if cell_data['border']: cell.border = cell_data['border']
            if cell_data['alignment']: cell.alignment = cell_data['alignment']
            if cell_data['number_format']: cell.number_format = cell_data['number_format']

def main():
    import openpyxl
    print("\n==========================================================")
    print("      ระบบแยกประเภทและจัดเก็บประวัติงานสำเร็จรูป (AVEAM)")
    print("==========================================================")
    print(f"กำลังวิเคราะห์ไฟล์งานหลัก: {os.path.basename(F5_PATH)}...")
    
    if not os.path.exists(F5_PATH):
        print(f"[Error] ไม่พบไฟล์ Excel ที่ {F5_PATH}")
        sys.exit(1)
        
    # Open workbook in value-mode and formula-mode
    wb_val = openpyxl.load_workbook(F5_PATH, data_only=True)
    wb_form = openpyxl.load_workbook(F5_PATH, data_only=False)
    
    sh_val = wb_val['PN_Job_Relationship']
    sh_form = wb_form['PN_Job_Relationship']
    
    completed_rows = []
    incomplete_rows = []
    
    # Read active file rows
    for r in range(2, sh_form.max_row + 1):
        cust_val = sh_val.cell(r, 2).value
        pn_val = sh_val.cell(r, 7).value
        # Skip completely empty rows
        if not cust_val and not pn_val:
            continue
            
        row_data = get_row_data(sh_val, sh_form, r)
        if row_data['is_completed']:
            completed_rows.append(row_data)
        else:
            incomplete_rows.append(row_data)
            
    print(f"วิเคราะห์แถวข้อมูลทั้งหมด: {len(completed_rows) + len(incomplete_rows)} แถว")
    print(f"  - งานที่ประกอบเสร็จสมบูรณ์แล้ว (Completed): {len(completed_rows)} จ๊อบ")
    print(f"  - งานที่ยังไม่เสร็จ (WIP / Incomplete): {len(incomplete_rows)} จ๊อบ")
    
    if not completed_rows:
        print("\n🎉 ไม่พบงานประกอบตัวใหม่ที่เสร็จสมบูรณ์แล้ว (ไม่มีการย้ายข้อมูล)")
        return
        
    # --- ARRANGE ARCHIVE FILE ---
    existing_completed = []
    
    if os.path.exists(COMPLETED_PATH):
        print(f"\nพบไฟล์เก็บประวัติเดิม: {os.path.basename(COMPLETED_PATH)}")
        print("กำลังอ่านข้อมูลประวัติเดิม...")
        wb_arch_val = openpyxl.load_workbook(COMPLETED_PATH, data_only=True)
        wb_arch_form = openpyxl.load_workbook(COMPLETED_PATH, data_only=False)
        sh_arch_val = wb_arch_val['PN_Job_Relationship']
        sh_arch_form = wb_arch_form['PN_Job_Relationship']
        
        for r in range(2, sh_arch_form.max_row + 1):
            cust_val = sh_arch_val.cell(r, 2).value
            pn_val = sh_arch_val.cell(r, 7).value
            if not cust_val and not pn_val:
                continue
                
            cells = []
            for col in range(1, sh_arch_form.max_column + 1):
                cell = sh_arch_form.cell(row=r, column=col)
                cells.append({
                    'value': cell.value,
                    'data_type': cell.data_type,
                    'font': copy.copy(cell.font),
                    'fill': copy.copy(cell.fill),
                    'border': copy.copy(cell.border),
                    'alignment': copy.copy(cell.alignment),
                    'number_format': cell.number_format
                })
            val_q = sh_arch_val.cell(r, 17).value
            existing_completed.append({
                'original_row_index': r,
                'cells': cells,
                'sort_val': val_q
            })
        print(f"  - โหลดประวัติงานเดิมสำเร็จ: {len(existing_completed)} รายการ")
    else:
        print(f"\nไม่พบไฟล์เก็บประวัติเดิม | กำลังสร้างไฟล์ใหม่: {os.path.basename(COMPLETED_PATH)} จากแม่แบบ...")
        shutil.copy2(F5_PATH, COMPLETED_PATH)
        
    # Merge existing history with newly completed rows
    all_completed = existing_completed + completed_rows
    
    # Sort completed rows by Column Q (Latest Ship Date)
    print("กำลังจัดเรียงลำดับงานประกอบเสร็จตาม Latest Ship Date (Column Q)...")
    all_completed.sort(key=lambda x: get_date_key(x['sort_val']))
    
    # Write to Completed File
    print(f"กำลังจัดเก็บรายการที่เสร็จแล้วทั้งหมด {len(all_completed)} แถวลงในไฟล์ประวัติ...")
    wb_completed = openpyxl.load_workbook(COMPLETED_PATH, data_only=False)
    sh_completed = wb_completed['PN_Job_Relationship']
    clear_rows(sh_completed)
    write_rows(sh_completed, all_completed)
    wb_completed.save(COMPLETED_PATH)
    print("บันทึกไฟล์ประวัติงานเสร็จสะสมสำเร็จ!")
    
    # --- WRITE ACTIVE INCOMPLETE ROWS BACK TO F5 ---
    print(f"\nกำลังอัปเดตไฟล์หลัก {os.path.basename(F5_PATH)} (ให้คงเหลือเฉพาะงาน WIP {len(incomplete_rows)} รายการ)...")
    clear_rows(sh_form)
    write_rows(sh_form, incomplete_rows)
    wb_form.save(F5_PATH)
    print("อัปเดตไฟล์หลักสำเร็จเรียบร้อย!")
    
    print("\n==========================================================")
    print("      ดำเนินโครงการเสร็จสมบูรณ์และย้ายประวัติเรียบร้อย! ✨")
    print(f"  - งานที่ยังไม่เสร็จ (WIP) ในไฟล์หลักคงเหลือ: {len(incomplete_rows)} รายการ")
    print(f"  - งานที่เสร็จสมบูรณ์สะสมในไฟล์ประวัติ: {len(all_completed)} รายการ")
    print("==========================================================")

if __name__ == '__main__':
    main()
